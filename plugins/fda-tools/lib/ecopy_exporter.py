"""
eCopy Exporter for FDA 510(k) Submissions

This module exports 510(k) submission packages in FDA eCopy format per
"eCopy Program for Medical Device Submissions" guidance (August 2019).

Features:
- FDA eCopy folder structure (0001-cover-letter through 0019-mri-safety)
- Markdown to PDF conversion using pandoc
- PDF styling: Times New Roman 12pt, 1-inch margins, TOC, numbered sections
- eCopy checklist generation (Excel)
- Size validation (<200 MB FDA limit per section)

Sprint 6 Feature 3: eCopy Export Command (91.3 → 92.8 expert rating)

Author: Claude Code (Anthropic)
Date: 2026-02-14
"""

import os
import sys
import json
import subprocess
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime


class eCopyExporter:
    """
    Export 510(k) submission package in FDA eCopy format.

    FDA eCopy Structure (per August 2019 guidance):
    - Numbered folders: 0001-xxx through 0019-xxx
    - PDF files only (with original source files as supplements)
    - eCopy-Checklist.xlsx manifest
    - Total package size <200 MB per section, <4 GB total
    """

    # FDA eCopy folder mapping
    ECOPY_SECTIONS = {
        "01": {"name": "Cover Letter", "draft_file": "draft_cover-letter.md"},
        "02": {
            "name": "Administrative",
            "draft_files": [
                "draft_truthful-accuracy.md",
                "draft_form-3514.md",
                "draft_financial-certification.md",
            ],
        },
        "03": {
            "name": "510k Summary",
            "draft_files": ["draft_form-3881.md", "draft_510k-summary.md"],
        },
        "04": {
            "name": "Device Description",
            "draft_files": ["draft_device-description.md"],
        },
        "05": {
            "name": "Substantial Equivalence",
            "draft_files": [
                "draft_se-discussion.md",
                "se_comparison.md",
                "draft_predicate-justification.md",
            ],
        },
        "06": {
            "name": "Proposed Labeling",
            "draft_files": [
                "draft_labeling.md",
                "IFU.md",
                "draft_IFU.md",
                "labels.md",
            ],
        },
        "07": {"name": "Sterilization", "draft_files": ["draft_sterilization.md"]},
        "08": {"name": "Biocompatibility", "draft_files": ["draft_biocompatibility.md"]},
        "09": {"name": "Software", "draft_files": ["draft_software.md"]},
        "10": {
            "name": "Electrical Safety EMC",
            "draft_files": ["draft_emc-electrical.md"],
        },
        "11": {"name": "Shelf Life", "draft_files": ["draft_shelf-life.md"]},
        "12": {"name": "Performance Testing", "draft_files": ["draft_performance-summary.md", "draft_testing-rationale.md", "test_plan.md"]},
        "13": {"name": "Clinical", "draft_files": ["draft_clinical.md"]},
        "14": {"name": "Human Factors", "draft_files": ["draft_human-factors.md"]},
        "15": {
            "name": "Combination Product",
            "draft_files": ["draft_combination-product.md"],
        },
        "16": {"name": "Reprocessing", "draft_files": ["draft_reprocessing.md"]},
        "17": {
            "name": "Declaration of Conformity",
            "draft_files": ["draft_doc.md", "standards_lookup.json"],
        },
        "18": {"name": "Other", "draft_files": []},
        "19": {"name": "MRI Safety", "draft_files": ["draft_mri-safety.md"]},  # NEW
    }

    def __init__(self, project_path: str):
        """
        Initialize eCopy exporter.

        Args:
            project_path: Path to 510(k) project directory
        """
        self.project_path = Path(project_path)
        self.ecopy_path = self.project_path / "eCopy"
        self.drafts_path = self.project_path / "drafts"

        # Check pandoc availability
        self.pandoc_available = self._check_pandoc()

    def _check_pandoc(self) -> bool:
        """
        Check if pandoc is installed.

        Returns:
            True if pandoc is available, False otherwise
        """
        try:
            result = subprocess.run(
                ["pandoc", "--version"],
                capture_output=True,
                text=True,
                timeout=5,
            )
            return result.returncode == 0
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError, Exception):
            # Gracefully handle any subprocess errors
            return False

    def export(self) -> Dict[str, Any]:
        """
        Export full eCopy package.

        Returns:
            Dict with keys:
                - ecopy_path (str): Path to eCopy directory
                - sections_created (int): Number of sections created
                - files_converted (int): Number of files converted to PDF
                - total_size_mb (float): Total package size in MB
                - validation (Dict): Validation results
        """
        # Create eCopy directory structure
        self.ecopy_path.mkdir(exist_ok=True)

        sections_created = 0
        files_converted = 0
        conversion_errors = []

        # Process each eCopy section
        for section_num, section_info in self.ECOPY_SECTIONS.items():
            section_name = section_info["name"]
            section_dir = self.ecopy_path / f"{section_num.zfill(4)}-{section_name.replace(' ', '')}"
            section_dir.mkdir(exist_ok=True)
            sections_created += 1

            # Find and convert draft files
            draft_files = section_info.get("draft_files", section_info.get("draft_file"))
            if isinstance(draft_files, str):
                draft_files = [draft_files]

            for draft_file in draft_files:
                draft_path = self.drafts_path / draft_file
                if not draft_path.exists():
                    draft_path = self.project_path / draft_file
                if not draft_path.exists():
                    continue

                # Convert markdown to PDF
                output_pdf = section_dir / f"{draft_path.stem}.pdf"

                if self.pandoc_available:
                    try:
                        self._convert_md_to_pdf(draft_path, output_pdf)
                        files_converted += 1
                    except Exception as e:
                        conversion_errors.append(
                            f"{draft_file}: {str(e)}"
                        )
                else:
                    # Copy markdown file if pandoc not available
                    import shutil
                    shutil.copy(draft_path, section_dir / draft_file)
                    files_converted += 1

        # Calculate total size
        total_size_mb = sum(
            f.stat().st_size for f in self.ecopy_path.rglob("*") if f.is_file()
        ) / (1024 * 1024)

        # Generate eCopy checklist
        checklist_path = self.ecopy_path / "eCopy-Checklist.xlsx"
        self._generate_ecopy_checklist(checklist_path)

        # Validation
        validation = self._validate_ecopy_package()

        return {
            "ecopy_path": str(self.ecopy_path),
            "sections_created": sections_created,
            "files_converted": files_converted,
            "total_size_mb": round(total_size_mb, 2),
            "validation": validation,
            "conversion_errors": conversion_errors,
            "pandoc_available": self.pandoc_available,
        }

    def _convert_md_to_pdf(self, input_md: Path, output_pdf: Path):
        """
        Convert markdown file to PDF using pandoc with FDA-compliant styling.

        Args:
            input_md: Path to input markdown file
            output_pdf: Path to output PDF file
        """
        # Pandoc command with FDA-compliant styling
        cmd = [
            "pandoc",
            str(input_md),
            "-o",
            str(output_pdf),
            "--pdf-engine=xelatex",
            "--toc",  # Table of contents
            "--number-sections",  # Numbered sections
            "-V", "mainfont=Times New Roman",
            "-V", "fontsize=12pt",
            "-V", "geometry:margin=1in",
            "-V", "linkcolor=blue",
            "-V", "toc-title=Table of Contents",
            "--metadata", f"date={datetime.now().strftime('%Y-%m-%d')}",
        ]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,
        )

        if result.returncode != 0:
            raise RuntimeError(f"Pandoc conversion failed: {result.stderr}")

    def _generate_ecopy_checklist(self, checklist_path: Path):
        """
        Generate eCopy checklist Excel file.

        Args:
            checklist_path: Path to output Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "eCopy Checklist"

            # Header row
            headers = [
                "Section #",
                "Section Title",
                "File Name",
                "File Size (KB)",
                "Status",
                "Notes",
            ]
            ws.append(headers)

            # Style header row
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Populate checklist
            row_num = 2
            for section_num, section_info in self.ECOPY_SECTIONS.items():
                section_name = section_info["name"]
                section_dir = (
                    self.ecopy_path
                    / f"{section_num.zfill(4)}-{section_name.replace(' ', '')}"
                )

                if section_dir.exists():
                    files = list(section_dir.glob("*.pdf")) + list(
                        section_dir.glob("*.md")
                    )
                    if files:
                        for file in files:
                            file_size_kb = file.stat().st_size / 1024
                            ws.append(
                                [
                                    section_num,
                                    section_name,
                                    file.name,
                                    round(file_size_kb, 2),
                                    "Complete",
                                    "",
                                ]
                            )
                            row_num += 1
                    else:
                        ws.append(
                            [section_num, section_name, "N/A", 0, "Empty", ""]
                        )
                        row_num += 1
                else:
                    ws.append(
                        [section_num, section_name, "N/A", 0, "Missing", ""]
                    )
                    row_num += 1

            # Adjust column widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 40
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 30

            wb.save(checklist_path)

        except ImportError:
            # openpyxl not available - create simple CSV instead
            with open(checklist_path.with_suffix(".csv"), "w") as f:
                f.write("Section #,Section Title,File Name,File Size (KB),Status,Notes\n")

                for section_num, section_info in self.ECOPY_SECTIONS.items():
                    section_name = section_info["name"]
                    section_dir = (
                        self.ecopy_path
                        / f"{section_num.zfill(4)}-{section_name.replace(' ', '')}"
                    )

                    if section_dir.exists():
                        files = list(section_dir.glob("*.pdf")) + list(
                            section_dir.glob("*.md")
                        )
                        if files:
                            for file in files:
                                file_size_kb = file.stat().st_size / 1024
                                f.write(
                                    f"{section_num},{section_name},{file.name},{round(file_size_kb, 2)},Complete,\n"
                                )
                        else:
                            f.write(f"{section_num},{section_name},N/A,0,Empty,\n")
                    else:
                        f.write(f"{section_num},{section_name},N/A,0,Missing,\n")

    def _validate_ecopy_package(self) -> Dict[str, Any]:
        """
        Validate eCopy package completeness and size limits.

        Returns:
            Dict with validation results
        """
        validation = {
            "status": "PASS",
            "mandatory_sections": True,
            "file_naming": True,
            "size_ok": True,
            "checklist_ok": True,
            "errors": [],
            "warnings": [],
        }

        # Check mandatory sections (01, 02, 03, 04, 05, 06)
        mandatory_sections = ["01", "02", "03", "04", "05", "06"]
        for section_num in mandatory_sections:
            section_info = self.ECOPY_SECTIONS[section_num]
            section_name = section_info["name"]
            section_dir = (
                self.ecopy_path
                / f"{section_num.zfill(4)}-{section_name.replace(' ', '')}"
            )

            if not section_dir.exists() or not list(section_dir.glob("*")):
                validation["mandatory_sections"] = False
                validation["errors"].append(
                    f"Mandatory section {section_num} ({section_name}) is empty or missing"
                )

        # Check total package size (<4 GB FDA limit)
        total_size_gb = sum(
            f.stat().st_size for f in self.ecopy_path.rglob("*") if f.is_file()
        ) / (1024**3)

        if total_size_gb >= 4.0:
            validation["size_ok"] = False
            validation["errors"].append(
                f"Package size {total_size_gb:.2f} GB exceeds FDA 4 GB limit"
            )
        elif total_size_gb >= 3.5:
            validation["warnings"].append(
                f"Package size {total_size_gb:.2f} GB approaching FDA 4 GB limit"
            )

        # Check individual section sizes (<200 MB recommended)
        for section_dir in self.ecopy_path.glob("????-*"):
            if section_dir.is_dir():
                section_size_mb = sum(
                    f.stat().st_size for f in section_dir.rglob("*") if f.is_file()
                ) / (1024**2)

                if section_size_mb >= 200:
                    validation["warnings"].append(
                        f"Section {section_dir.name} size {section_size_mb:.2f} MB exceeds recommended 200 MB"
                    )

        # Check checklist exists
        if not (self.ecopy_path / "eCopy-Checklist.xlsx").exists() and not (
            self.ecopy_path / "eCopy-Checklist.csv"
        ).exists():
            validation["checklist_ok"] = False
            validation["errors"].append("eCopy checklist not generated")

        # Set overall status
        if validation["errors"]:
            validation["status"] = "FAIL"
        elif validation["warnings"]:
            validation["status"] = "WARNING"

        return validation


def export_ecopy(project_path: str) -> Dict[str, Any]:
    """
    Convenience function to export eCopy package.

    Args:
        project_path: Path to 510(k) project directory

    Returns:
        Export results dict
    """
    exporter = eCopyExporter(project_path)
    return exporter.export()
